import openpyxl
import json

excel_file = "SSJ - Prime Database TAD.xlsx"
sheet_name = "SSJ (4710002085) 2026.05.26 " # atau gunakan wb.active

print(f"Membaca file Database Primer: {excel_file}...")

try:
    wb = openpyxl.load_workbook(excel_file, data_only=True)
    
    # Mengambil active sheet secara otomatis
    sheet = wb.active

    # Membaca header dari Baris 6 (total 73 kolom)
    headers = [sheet.cell(row=6, column=c).value for c in range(1, 74)]

    employee_list = []

    # Membaca data karyawan dari Baris 7 sampai 58 (52 baris data)
    for row in range(7, 59):
        no_val = sheet.cell(row=row, column=1).value
        if no_val is None:
            continue
            
        emp_data = {}
        for c in range(1, 74):
            h = headers[c-1]
            if h is not None:
                val = sheet.cell(row=row, column=c).value
                # Ubah format tanggal jika ada agar terbaca dengan baik di JSON
                if hasattr(val, 'strftime'):
                    val = val.strftime('%Y-%m-%d')
                emp_data[str(h)] = val if val is not None else ""
        
        employee_list.append(emp_data)

    output_json = "employee_data.json"
    with open(output_json, "w", encoding="utf-8") as f:
        json.dump(employee_list, f, indent=4, ensure_ascii=False)

    print(f"SUKSES! Berhasil memperbarui {len(employee_list)} data karyawan primer ke {output_json}.")

except Exception as e:
    print(f"Terjadi kesalahan saat memproses Database Primer: {e}")